from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import load_workbook
from .models import State_Master  


def import_from_excel(request):
    # print("outside condition")
    if request.method=='GET':
        # print("inside condition")
        file= 'C:\\Users\\Lenovo\\Desktop\\PM\\state_data.xlsx'
        wb = load_workbook(file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            name_state,code_state= row
            obj=State_Master(state_name=name_state,state_code=code_state)
            obj.save()

    return HttpResponse("done")

